from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font


def _write_sheet(wb, title, headers, rows):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append(row)
    for i, header in enumerate(headers, start=1):
        width = max(len(header) + 2, 20)
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(width + 20, 60)
    return ws


def build_workbook(results: dict) -> BytesIO:
    wb = Workbook()
    wb.remove(wb.active)

    risk_rows = [
        (
            r.get("id", ""),
            r.get("category", ""),
            r.get("description", ""),
            "\n".join(r.get("suggested_controls", [])),
        )
        for r in results.get("risks", [])
    ]
    _write_sheet(
        wb,
        "Risk Register",
        ["ID", "Category", "Risk Description", "Suggested Controls"],
        risk_rows,
    )

    opp_rows = [
        (
            o.get("id", ""),
            o.get("description", ""),
            "\n".join(o.get("suggested_actions", [])),
        )
        for o in results.get("opportunities", [])
    ]
    _write_sheet(
        wb,
        "Opportunities",
        ["ID", "Opportunity Description", "Suggested Actions"],
        opp_rows,
    )

    question_rows = [(q,) for q in results.get("probing_questions", [])]
    _write_sheet(wb, "Probing Questions", ["Question"], question_rows)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
